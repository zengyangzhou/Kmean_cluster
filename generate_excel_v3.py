import openpyxl
import datetime
import sys
import numpy as np
from sklearn.preprocessing import Normalizer
# start_time = datetime.datetime.now()
# f = openpyxl.Workbook()
# sheet1 = f.active
# for i in range(1, 2000):
#     for k in range(1, 500):
#         sheet1.cell(row=i, column=k, value='asd')ls
# f.save('test_of_openpyxl.xlsx')
# end_time = datetime.datetime.now()
# print('Time Consumption', end_time - start_time)
'''
version3: It can calculate average cluster distance to centroid of each cluster and distance to centroid of each example, 同时排序

python generate_excel_v3.py '/web/home/zengyangzhou/bert/cluster_sentence_7_5000_200.npy' '/web/home/zengyangzhou/data/chat/chat_20190305.txt' '/web/home/zengyangzhou/bert/cluster_centroids_7_5000_200.npy' 200 'excel_cluster_7_5000_200.xlsx' '/web/home/zengyangzhou/bert/sentence_embedding_list_7_5000.npy' '/web/home/zengyangzhou/bert/cluster_labels_7_5000_200.npy'

python generate_excel_v3.py '/web/home/zengyangzhou/bert/cluster_sentence_7_v300.npy' '/web/home/zengyangzhou/data/chat/chat_20190305.txt' '/web/home/zengyangzhou/bert/cluster_centroids_7_v300.npy' 300 'excel_cluster_7_v300.xlsx' '/web/home/zengyangzhou/bert/sentence_embedding_list_7.npy' '/web/home/zengyangzhou/bert/cluster_labels_7_v300.npy'

python generate_excel_v3.py '/web/home/zengyangzhou/bert/cluster_sentence_7_v150.npy' '/web/home/zengyangzhou/data/chat/chat_20190305.txt' '/web/home/zengyangzhou/bert/cluster_centroids_7_v150.npy' 150 'excel_cluster_7_v150.xlsx' '/web/home/zengyangzhou/bert/sentence_embedding_list_7.npy' '/web/home/zengyangzhou/bert/cluster_labels_7_v150.npy'


python generate_excel_v3.py '/web/home/zengyangzhou/bert/cluster_sentence_63.npy' '/web/home/zengyangzhou/bert/data.txt' '/web/home/zengyangzhou/bert/cluster_centroids_63.npy' 10 'excel_cluster_63.xlsx' '/web/home/zengyangzhou/bert/sentence_embedding_list_63.npy' '/web/home/zengyangzhou/bert/cluster_labels_63.npy'

python generate_excel_v3.py '/web/home/zengyangzhou/bert/task_2/cluster_sentence_data2_v10.npy' '/web/home/zengyangzhou/bert/task_2/data_2.txt' '/web/home/zengyangzhou/bert/task_2/cluster_centroids_data_2_v10.npy' 10 '/web/home/zengyangzhou/bert/task_2/excel_cluster_data_2_v10.xlsx' '/web/home/zengyangzhou/bert/task_2/sentence_embedding_list_data_2.npy' '/web/home/zengyangzhou/bert/task_2/cluster_labels_data_2_v10.npy'

python generate_excel_v3.py '/web/home/zengyangzhou/bert/task_2/cluster_sentence_data2_v20.npy' '/web/home/zengyangzhou/bert/task_2/data_2.txt' '/web/home/zengyangzhou/bert/task_2/cluster_centroids_data_2_v20.npy' 20 '/web/home/zengyangzhou/bert/task_2/excel_cluster_data_2_v20.xlsx' '/web/home/zengyangzhou/bert/task_2/sentence_embedding_list_data_2.npy' '/web/home/zengyangzhou/bert/task_2/cluster_labels_data_2_v20.npy'

'''
def get_key_of_dict(dict, value):
    return list([k for k, v in dict.items() if v == value])

def generate_excel_from_cluster(cluster_sentence_path, chat_data_path,
                                cluster_centroid_file, cluster_num,
                                result_cluster_excel_output,
                                sentence_embedding_file, cluster_labels_file):
    # with open ('/web/home/zengyangzhou/data/chat/chat_20190305.txt', 'r', encoding='utf-8') as file:
    with open(chat_data_path, 'r', encoding='utf-8') as file:
        data = []
        for line in file.readlines():
            line = line.strip().replace(' ', '')
            data.append(line)

    #data = data[:train_size]  # 只取数据的前train_size个用于测试不同的 cluster_num 的效果
    the_cluster = np.load(cluster_sentence_path)

    the_centroids = np.load(cluster_centroid_file)

    # 归一化句向量
    sentence_embedding_list = np.load(sentence_embedding_file)
    normalizer = Normalizer(copy=False)
    sentence_embedding_list_norm = normalizer.fit_transform(sentence_embedding_list)

    train_size = int(np.shape(the_cluster)[0]) # 只取数据的前train_size个用于测试不同的 cluster_num 的效果
    data = data[:train_size]
    check_dic = {}
    i = 0
    if np.shape(the_cluster) == np.shape(data):
        for cluster_index in the_cluster:
            check_dic[data[i]] = cluster_index
            i = i + 1
        # print('check_dic', check_dic)
    else:
        print('ERROR: shapes are not right')
        print('shape of the_cluster', np.shape(the_cluster))
        print('shape of data', np.shape(data))
    #performance_of_cluster = open(output_file_path, 'w')
    cluster_num = int(cluster_num)

    # creat excel ---------------------------------------------------------------------------
    f = openpyxl.Workbook()
    sheet1 = f.active
    sheet1.cell(row=1, column=1, value='Cluster_type')
    sheet1.cell(row=2, column=1, value='Number of Query')
    sheet1.cell(row=3, column=1, value='Topic')
    sheet1.cell(row=4, column=1, value='Cluster_Performance(Cluster Average Distance)')
    column_number = 2
    labels = np.load(cluster_labels_file)
    #lables 是一个list每个值是对应query的分类序号
    average_distance_query_list = []
    num_query_cluster_list = []
    query_order_list = []
    average_distance_cluster_list = []
    for cluster_order in range(cluster_num): #cluster_order也就是聚类的类序号
        #keys = get_key_of_dict(check_dic, cluster_order)

        sheet1.cell(row=1, column=column_number, value=cluster_order)
        #sheet1.cell(row=2, column=column_number, value=len(keys))

        sheet1.cell(row=3, column=column_number, value='') #给TOPIC留的空位
        #sheet1.cell(row=4, column=column_number, value='') #数值为类内所有点的平均距离，还没有算

        sheet1.cell(row=1, column=column_number + 1, value='Distance to Centroid(Each Query)')
        sheet1.cell(row=2, column=column_number + 1, value='/')
        sheet1.cell(row=3, column=column_number + 1, value='/')
        sheet1.cell(row=4, column=column_number + 1, value='/')

        sheet1.cell(row=1, column=column_number + 2, value='Label')
        sheet1.cell(row=2, column=column_number + 2, value='/')
        sheet1.cell(row=3, column=column_number + 2, value='/')
        sheet1.cell(row=4, column=column_number + 2, value='/')

        #sheet1.cell(row=row_number + 5, column=column_number, value=str(keys[row_number]))
        #计算每个Query到中心的距离,找到分类的序号和每个Query在data中序号的对应关系
        locations_of_label = np.argwhere(labels == cluster_order)
        #是一个list
        sheet1.cell(row=2, column=column_number, value=len(locations_of_label))
        #生成每一类对应的query在data中的对应位置
        average_distance_cluster = 0
        row_number = 0
        for query_order in locations_of_label:
            #query_order 就是query在整个chatdata中的序号
            query_order = int(query_order)
            average_distance_query = \
                np.linalg.norm(the_centroids[cluster_order] - sentence_embedding_list_norm[query_order])
            #sheet1.cell(row=row_number + 5, column=column_number, value=data[query_order])
            #sheet1.cell(row=row_number + 5, column=column_number + 1, value=average_distance_query)
            average_distance_cluster = average_distance_cluster + average_distance_query
            average_distance_query_list.append(average_distance_query)
            query_order_list.append(query_order)
            row_number = row_number + 1
        if len(locations_of_label) > 1:
            average_distance_cluster = average_distance_cluster/len(locations_of_label)
        if len(locations_of_label) == 1:
            average_distance_cluster = average_distance_cluster
        #sheet1.cell(row=4, column=column_number, value=average_distance_cluster)
        average_distance_cluster_list.append(average_distance_cluster)

        #统计每一类query的数量
        num_query_cluster_list.append(len(locations_of_label))
        column_number = column_number + 3
    #表格的数据已经完成
    #print('average_distance_cluster_list', average_distance_cluster_list)
    #print('\n')
    #print('average_distance_query_list', average_distance_query_list)
    average_distance_query_list = np.array(average_distance_query_list)
    average_distance_query_list = average_distance_query_list[:, np.newaxis]
    average_distance_query_list = list(average_distance_query_list)
    num_query_cluster_list = np.array(num_query_cluster_list)
    num_query_cluster_list = num_query_cluster_list[:, np.newaxis]
    query_order_list = np.array(query_order_list)
    query_order_list = query_order_list[:, np.newaxis]
    average_distance_cluster_list = np.array(average_distance_cluster_list)
    average_distance_cluster_list = average_distance_cluster_list[:, np.newaxis]
    average_distance_cluster_list = list(average_distance_cluster_list)
    #print('average_distance_query_list', np.shape(average_distance_query_list))
    # 每个query到中心的距离，尺寸为（train_size，）
    #print('query_order_list', np.shape(query_order_list))
    #print(type(query_order_list[100]))
    # 每个query在data中的对应序号，从1到109359，尺寸为（train_size，）
    #print('average_distance_cluster_list', np.shape(average_distance_cluster_list))
    #print(average_distance_cluster_list[50])
    #print(type(average_distance_cluster_list[50]))
    # 每个cluster到中心的平均距离，尺寸为（cluster_num，）
    #print('num_query_cluster_list', np.shape(num_query_cluster_list))
    # 每一类中query的数目，尺寸为（cluster_num，）
    #print('train_size and num_query_cluster_list', train_size, np.sum(num_query_cluster_list))
    # 检测是否每一类的query数目统计正确

    m = 0

    dict_ordered_query_dist_and_data = [] #query数目 x 1
    dict_average_dist_and_query_dist = {}
    dict_ordered_average_dist_and_query_dist = {} # query数目 x 1
    print('train_size', train_size)
    u = 0
    for cluster_order2 in range(cluster_num):
        dict_query_dist_and_data = {}
        dict_ordered_query_dist_and_data_each_cluster = []
        for i in range(int(num_query_cluster_list[cluster_order2])):
            if i <= 9:
                dict_query_dist_and_data[str(i) + ' ||| ' + data[int(query_order_list[u])]] = float(average_distance_query_list[u])
            else:
                dict_query_dist_and_data[str(i) + '||| ' + data[int(query_order_list[u])]] = float(average_distance_query_list[u])

            u = u + 1
        #print('dict_query_dist_and_data', dict_query_dist_and_data)
        #print('\n')
        dict_ordered_query_dist_and_data_each_cluster = \
            sorted(dict_query_dist_and_data.items(), key=lambda item: item[1])
        dict_ordered_query_dist_and_data.append(dict_ordered_query_dist_and_data_each_cluster)
    print('u == train_size', u == train_size)
    #print('shape of dict_ordered_query_dist_and_data', np.shape(dict_ordered_query_dist_and_data))

    #print('dict_ordered_query_dist_and_data', dict_ordered_query_dist_and_data)
    # print(dict_ordered_query_dist_and_data[0][0][0])
    # print(dict_ordered_query_dist_and_data[0][0])
    # print(dict_ordered_query_dist_and_data[0])
    #对与每一类的query进行排序，而不是一起全部排序
    # print('i', i)  #4999
    # print('shape of data', np.shape(data)) #(5000,)
    # print('shape of average_distance_query_list', np.shape(average_distance_query_list))  #(5000,1)
    #print('dict_query_dist_and_data', np.shape(dict_query_dist_and_data))
    #dict_ordered_query_dist_and_data = sorted(dict_query_dist_and_data.items(), key=lambda item: item[1])
    #type = list, 其中的元素是tuple【（‘我在干嘛’，0.565），...】
    #形成了key是对应的query， value是对应query到中心距离的有序字典
    #print('dict_ordered_query_dist_and_data', len(dict_ordered_query_dist_and_data))
    # print('length of dict_query_dist_and_data', len(dict_query_dist_and_data))
    # print('shape of dict_ordered_query_dist_and_data', np.shape(dict_ordered_query_dist_and_data))
    #print(dict_query_dist_and_data[data[int(query_order_list[0])]], dict_query_dist_and_data[data[int(query_order_list[4999])]])
    #print(dict_ordered_query_dist_and_data[0], dict_ordered_query_dist_and_data[4998])
    #print(dict_ordered_query_dist_and_data[0], dict_ordered_query_dist_and_data[4999])
    # print(np.shape(average_distance_cluster_list))
    # print(dict_ordered_query_dist_and_data[4999])
    dict_cluster_query_number_and_average_dist = {}
    #这是每一类对应有多少个query，并且根据类平均距离排序后的字典
    for k in range(cluster_num):# 一共多少类
        #print(k)
        dict_cluster_query_number_and_average_dist[float(average_distance_cluster_list[k])] = int(num_query_cluster_list[k])
        for n in range(int(num_query_cluster_list[k])): #每一类中query数目
            #print('k', k)
            #print('m', m)

            dict_average_dist_and_query_dist[dict_ordered_query_dist_and_data[k][n]] = \
                float(average_distance_cluster_list[k])
            m = m + 1

    #print('m should = train_size', m)
    #print('shape of dict_average_dist_and_query_dist', len(dict_average_dist_and_query_dist))
    dict_ordered_average_dist_and_query_dist = \
        sorted(dict_average_dist_and_query_dist.items(), key=lambda item: item[1])
    list_cluster_query_number_and_average_dist = \
        sorted(dict_cluster_query_number_and_average_dist.items(), key=lambda item: item[0])
    #print('dict_ordered_average_dist_and_query_dist', dict_ordered_average_dist_and_query_dist)
    #print('list_cluster_query_number_and_average_dist', list_cluster_query_number_and_average_dist[0][1])
    #print('shape of dict_ordered_average_dist_and_query_dist', np.shape(dict_ordered_average_dist_and_query_dist))
    #形成了key是query和query到中心距离， value是类平均到中心距离的字典
    #{（‘我在看书’，0.56）：0.5，....}
    column_number = 2
    end_query_loc = 0
    #print(len(dict_ordered_average_dist_and_query_dist))
    for cluster_order in range(cluster_num): #类序号0-200
        #keys = get_key_of_dict(check_dic, cluster_order)
        row_number = 0
        #还没把不同的类分开

        sheet1.cell(row=4, column=column_number,
                    value=dict_ordered_average_dist_and_query_dist[end_query_loc][1])
                    #数值为类内所有点的平均距离
        # sheet1.cell(row=4, column=column_number,
        #             value=int(list_cluster_query_number_and_average_dist[cluster_order][0]))
        # # 数值为类内所有点的平均距离
        sheet1.cell(row=2, column=column_number,
                    value=int(list_cluster_query_number_and_average_dist[cluster_order][1]))
        # 数值每一类中query的数量
        for query_order_in_cluster in range(int(list_cluster_query_number_and_average_dist[cluster_order][1])): #把每一类中的每一个QUERY进行操作
            #print('end_query_loc', end_query_loc)
            # if end_query_loc >= train_size - 1:
            #     #print('end_query_loc_max', end_query_loc)
            #     break
            sheet1.cell(row=row_number + 5 + query_order_in_cluster, column=column_number,
                        value=dict_ordered_average_dist_and_query_dist[end_query_loc][0][0]) #query
            sheet1.cell(row=row_number + 5 + query_order_in_cluster, column=column_number + 1,
                        value=dict_ordered_average_dist_and_query_dist[end_query_loc][0][1]) #query对应的距离
            end_query_loc += 1

        column_number = column_number + 3
    #print('end_query_loc == train_size', end_query_loc == train_size)
    f.save(result_cluster_excel_output)


if __name__ == '__main__':
    start_time = datetime.datetime.now()
    cluster_sentence_path = sys.argv[1]
    chat_data_path = sys.argv[2]
    cluster_centroid_file = sys.argv[3]
    #output_file_path = sys.argv[3]
    cluster_num = sys.argv[4]
    #train_size = sys.argv[4]
    result_cluster_excel_output = sys.argv[5]
    sentence_embedding_file = sys.argv[6]
    cluster_labels_file = sys.argv[7]
    if len(sys.argv) < 8:
        print('1: cluster_sentence_path 2: chat_data_path, 3: cluster_num , 4: result_cluster_output_path')
    elif len(sys.argv) == 8:
        generate_excel_from_cluster(cluster_sentence_path, chat_data_path, cluster_centroid_file, cluster_num, result_cluster_excel_output, sentence_embedding_file, cluster_labels_file)
    else:
        print('way of input is wrong')
        print(len(sys.argv))
    end_time = datetime.datetime.now()
    print('Time Cost', end_time - start_time)
